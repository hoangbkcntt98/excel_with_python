
import pandas as pd
from src.backend.config import  header
import os
import os.path
from os import path
import csv
import numpy as np
from openpyxl import load_workbook
from src.util.util_string import remove_nan
from src.config import __output_dir__
def fill_title(data_xls):
    for i in range(len(data_xls)):
        title_no = data_xls.loc[i, header['title_no']]
        title_name = data_xls.loc[i, header['title_name']]
        if pd.isna(title_no) or pd.isna(title_name):
            title_no = old_title_no
            title_name  = old_title_name
        else:
            old_title_no = title_no
            old_title_name = title_name
        data_xls.loc[i, header['title_no']] = title_no
        data_xls.loc[i, header['title_name']] = title_name
def get_last_row(data_xls, flg_title = 'sentence'):
    for i in range(len(data_xls)):
        flag = data_xls.loc[i, header[flg_title]]
        if pd.isna(flag):
            return i
    return -1
def get_headers(dataframe):
    return list(dataframe.columns)

def read_execl_file(path, sheet = "K", reading_cols = "B:K", unness_rows= False, is_fill_title = False, is_last = False):
    if not unness_rows:
        unness_rows = None
    else:
        unness_rows = lambda x: x in unness_rows
    if reading_cols:
        data_xls = pd.read_excel(path, sheet, index_col=None, usecols=reading_cols, skiprows=unness_rows)
    else:
        data_xls = pd.read_excel(path, sheet, index_col=None, skiprows=unness_rows)
    if is_fill_title:
        fill_title(data_xls)
    headers = get_headers(data_xls)
    if is_last:
        last_row = get_last_row(data_xls)
        data_xls = data_xls.head(last_row)
    data = []

    for i in range(len(data_xls)):
        row = []
        [row.append(data_xls.loc[i, headers[j]]) for j in range(len(headers))]
        row = remove_nan(row)
        data.append(tuple(row))
    results = {
        'data': data,
        'header':headers,
        'data_frame': data_xls
    }
    return results
def get_excel_info(path_dir):
    wb = load_workbook(path_dir)
    result  = {
        'file_path': path_dir,
        'sheet_name': wb.sheetnames
    }
    print(wb.sheetnames)
    return result

def to_csv(new_file_path,data_xls, is_save = True):
    data_xls.to_csv(new_file_path, encoding='utf_8', na_rep='Na', index=False, float_format = "%d")
    if not is_save:
        os.remove(new_file_path)
def check_exist_file(path_dir):
    exist_file = str(path.exists(path_dir))
    if exist_file == "False":
        return False
    return True

def read_data_from_csv(path_dir, ordered_columns = False, deleted_columns = False, limit= False, skip= False):
    if not check_exist_file(path_dir):
        return {
            'header': [],
            'data': []
        }
    if skip:
        data_csv = pd.read_csv(path_dir, skiprows=skip)
    else:
        data_csv = pd.read_csv(path_dir)
    headers = get_headers(data_csv)
    if ordered_columns:
        data_csv = data_csv[ordered_columns]
        headers = ordered_columns
    if deleted_columns:
        headers = [x for x in headers if not x in deleted_columns]
    data = []

    for i in range(len(data_csv)):
        row = []
        if deleted_columns :
            [row.append(data_csv.loc[i, headers[j]]) for j in range(len(headers)) if not headers[j] in deleted_columns]
        else:
            [row.append(data_csv.loc[i, headers[j]]) for j in range(len(headers))]
        row = remove_nan(row)

        data.append(row)
    if limit:
        temp= []
        for i in range(len(data)):
            if i>=limit[0] and i<=limit[1]:
                temp.append(data[i])
        data = temp
    results = {
        'data': data,
        'header': headers,
    }
    return results


# def save_csv(path_dir,tree):
#     with open(path_dir, "w", newline='') as myfile:
#         csvwriter = csv.writer(myfile, delimiter=',')
#
#         for row_id in tree.get_children():
#             row = tree.item(row_id)['values']
#             print('save row:', row)
#             csvwriter.writerow(row)
def create_csv_from_list(path_dir, data):
    with open(path_dir, "w", newline='',encoding='utf_8') as myfile:
        csvwriter = csv.writer(myfile, delimiter=',')
        for row in data:
            csvwriter.writerow(row)

def get_sheet_name():
    file = __output_dir__+'file_info.csv'
    if not check_exist_file(file):
        return ""
    data = pd.read_csv(file, header=None)
    print(data)
    return data.loc[0,0]